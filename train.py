
# 这个是python当中让print都以python3的形式进行print,即把print视为函数
from __future__ import print_function
# 使得我们能够手动输入命令行参数
import argparse
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torchvision import datasets, transforms
from torch.autograd import Variable
from patch1dataloader import train_loader, val_loader
import os
from mymodel import Net
import openpyxl
import pandas as pd

#创建train_acc.csv和var_acc.csv文件，记录loss和accuracy
df = pd.DataFrame(columns=['step','train Loss','test Loss'])#列名
df.to_csv("C:/D/BASENet/DeepSRQ-master/CVIU_result/FRIQA_GMDConv_G8_128_300.csv",index=False) #路径可以根据需要更改


os.environ['CUDA_VISIBLE_DEVICES'] = "0"    #设置当前使用的GPU设备仅为0号设备 设备名称为‘/gpu:0'
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
print(device)

# Training settings
# 设置一些参数,每个都有默认值,输入python main.py -h可以获得帮助
parser = argparse.ArgumentParser(description='Pytorch MNIST Example')
parser.add_argument('--batch-size', type=int, default=128, metavar='N',
                    help='input batch size for training (default: 64)')
parser.add_argument('--test-batch-size', type=int, default=1000, metavar='N',
                    help='input batch size for testing (default: 1000)')
parser.add_argument('--epochs', type=int, default=300, metavar='N',
                    help='number of epochs to train (default: 10')
parser.add_argument('--lr', type=float, default=0.01, metavar='LR',
                    help='learning rate (default: 0.01)')
parser.add_argument('--momentum', type=float, default=0.9, metavar='M',
                    help='SGD momentum (default: 0.9)')
parser.add_argument('--no-cuda', action='store_true', default=False,
                    help='disables CUDA training')
parser.add_argument('--seed', type=int, default=1, metavar='S',
                    help='random seed (default: 1)')
# 跑多少次batch进行一次日志记录
parser.add_argument('--log-interval', type=int, default=10, metavar='N',
                    help='how many batches to wait before logging training status')
# 这个是使用argparse模块时的必备行,将参数进行关联
args = parser.parse_args()
# 这个是在确认是否使用GPU的参数
args.cuda = not args.no_cuda and torch.cuda.is_available()

# 设置一个随机数种子
torch.manual_seed(args.seed)
if args.cuda:
    # 为GPU设置一个随机数种子
    torch.cuda.manual_seed(args.seed)

#train_loader = train_loader(batch_size = args.batch_size)
#val_loader = val_loader(batch_size = args.test_batch_size)
#train_set = datasets.MNIST(root='../dataset/mnist', train=True, transform=transform, download=False)
#train_loader = torch.utils.data.DataLoader(train_set, batch_size=args.batch_size, shuffle=True)
#test_set = datasets.MNIST(root='../dataset/mnist', train=False, transform=transform, download=False)
#test_loader = torch.utils.data.DataLoader(test_set, batch_size=args.test_batch_size, shuffle=False)
workbook = openpyxl.Workbook()
sheet0 = workbook.create_sheet(index=0) # 创建sheet0
sheet0.column_dimensions['A'].width=15 # 设置A列宽度
mymodel = Net()
#resume = '/home/liyx/DeepSRQ-master/CVIU_result/mymodelpatch1_30821281000.pth'
#checkpoint = torch.load(resume)
#mymodel.load_state_dict (checkpoint)
# 判断是否调用GPU模式
if args.cuda:
    mymodel.cuda()
# 初始化优化器 model.train()
optimizer = optim.SGD(mymodel.parameters(), lr=args.lr, weight_decay=1e-6,momentum=args.momentum,nesterov=True)
#optimizer = optim.Adam(mymodel.parameters(),lr=0.001,betas=(0.9, 0.999),eps=1e-08,weight_decay=1e-6,amsgrad=False)
mse_loss = torch.nn.MSELoss()

def train(epoch):
    """
    定义每个epoch的训练细节
    """
    # 设置为training模式
    mymodel.train()
    for batch_idx, (data1,data2, target) in enumerate(train_loader):
        # 如果要调用GPU模式,就把数据转存到GPU
        if args.cuda:
            data1,data2, target = data1.cuda(), data2.cuda(),target.cuda()
        data1, data2, target = Variable(data1), Variable(data2),Variable(target)

        # 优化器梯度初始化为零
        optimizer.zero_grad()
        output = mymodel(data1,data2)
        # 负对数似然函数损失
        loss = mse_loss(output.float(), target.float())
        loss.backward()
        optimizer.step()
        if batch_idx % args.log_interval == 0:
            print('Train Epoch: {} [{}/{} ({:.0f}%)]\tloss: {:.6f}'.format(
                epoch, batch_idx * len(data1), len(train_loader.dataset),
                100. * batch_idx / len(train_loader), loss.item()
            ))
    return loss


def test():
    # 设置为test模式
    mymodel.eval()
    # 初始化测试损失值为0
    test_loss = 0
    # 初始化预测正确的数据个数为0
    correct = 0
    int=0
    row = 0
    for data1,data2, target in val_loader:
        if args.cuda:
            data1,data2, target = data1.cuda(),data2.cuda(), target.cuda()
        data1,data2, target = Variable(data1), Variable(data2),Variable(target)
        output = mymodel(data1,data2)
        # 把所有loss值进行累加
        print(test_loss)
        int+=1
        test_loss += mse_loss(output.float(), target.float()).item()
        for i in range(len(output)):  # 逐行
            # print(i)
            # print(output[i].item())
            sheet0.cell(i + row + 1, 2).value = target[i].item()
            sheet0.cell(i + row + 1, 1).value = output[i].item()
            # sheet.write(i+row, 0,output[i].item())  # 将指定值写入第i行第j列
        row += len(output)
        # 获取最大对数概率值的索引
        #pred = output.data.max(1, keepdim=True)[1]
        # 对预测正确的个数进行累加
        #correct += pred.eq(target.data.view_as(pred)).cpu().sum()

    # 因为把所有loss值进行累加,所以最后要除以总的数据长度才能得到平均loss
    #print(test_loss)
    #print(len(val_loader.dataset))
    test_loss /= int
    workbook.save('./FRIQA_GMDConv_G8_128_300.xlsx')
    print('\nTest set: Average loss: {:.8f}, Accuracy: {}/{} ({:.0f}%)\n'.format(
        test_loss, correct, len(val_loader.dataset), 100. * correct / len(val_loader.dataset)
    ))
    return test_loss

train_loss = 0
test_loss = 0
best = 0
# 进行每个epoch的训练
for epoch in range(1, args.epochs + 1):
    step = "Step[%d]" % epoch
    t1_loss = train(epoch)
    train_loss = "%.8f" % t1_loss
    test()
    t2_loss = test()
    test_loss = "%.8f" % t2_loss
    list = [step, train_loss, test_loss]
    # 由于DataFrame是Pandas库中的一种数据结构，它类似excel，是一种二维表，所以需要将list以二维列表的形式转化为DataFrame
    data = pd.DataFrame([list])
    data.to_csv('C:/D/BASENet/DeepSRQ-master/CVIU_result/FRIQA_GMDConv_G8_128_300.csv', mode='a', header=False,index=False)  # mode设为a,就可以向csv文件追加数据了
    if epoch == 1:
     best = t2_loss
    elif best>t2_loss:
     best = t2_loss
     torch.save(mymodel.state_dict(),'C:/D/BASENet/DeepSRQ-master/CVIU_result/FRIQA_GMDConv_G8_128_300.pth')#保存训练PTH
